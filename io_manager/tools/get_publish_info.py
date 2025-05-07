from openpyxl import load_workbook
import os

def get_publish_info(xlsx_file_path):
    if os.path.exists(xlsx_file_path):
        wb = load_workbook(xlsx_file_path)
        ws = wb.active

        # 첫 번째 행에서 헤더 인덱스 찾기
        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
        headers = {}
        for col_idx, cell in enumerate(ws[1]):
            headers[cell.value] = col_idx

        # 각 행에서 dict 구성
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            seq = row[headers["seq_name"]]
            shot = row[headers["shot_name"]]
            typ = row[headers["type"]]
            directory = row[headers["Directory"]]
            if not (seq and shot and typ and directory):
                print(f"[SKIP] Not enough information in excel file : row num {row_idx}")
                continue  # 불완전한 행은 일단 건너뜀

            data.append({
                "sequence": str(seq),
                "shot": str(shot),
                "type": "org",
                "version": "v001",  # 고정
                "directory" : str(directory)
            })
        return data
    else:
        print(f"Not valid path : {xlsx_file_path} ")
        return