import subprocess
import json
import pandas as pd
import os
from PySide6.QtWidgets import QMessageBox
from tools.convert import exr_to_jpg, mov_to_jpg
import pyseq
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def export_metadata(date_path):
    # meta data 리스트, xlsx 파일의 한 행이 됌
    meta_list = []

    # 썸네일 저장 폴더 생성
    thumbnails_dir = os.path.join(date_path, "thumbnails")
    os.makedirs(thumbnails_dir, exist_ok=True)

    # .../scan/{date_path} 순회 하면서 Sequence 객체 get
    for root, dirs, seqs in pyseq.walk(date_path):
        dirs[:] = [d for d in dirs if d != "thumbnails"]
        scan_data_path = "" #exr의 첫 프레임 path / mov path
        for seq in seqs:
            _, ext = os.path.splitext(seq.name)
            if ext == ".xlsx":
                existing_xlsx_path = os.path.join(root, seq.name)
                print(f"[OK] Existing xlsx found: {existing_xlsx_path}")
                return existing_xlsx_path
            # EXR sequnece 폴더의 첫 프레임으로 thumnail(JPG) 추출
            elif ext == ".exr":
                scan_data_path = seq[0].path
                thumb_name = seq.head().strip(".") + ".jpg"
                thumb_path = os.path.join(thumbnails_dir, thumb_name)
                print(f"thumb path : {thumb_path} ")
                if exr_to_jpg(scan_data_path, thumb_path):
                    print(f"[OK] Thumbnail created: {thumb_path}")
                else:
                    print(f"[FAIL] Thumbnail create failed: {scan_data_path}")
                    thumb_path = ""
            # MOV 폴더의 첫 프레임 thumbnial(JPG) 추출
            elif ext == ".mov":
                mov_path = os.path.join(root, seq.name)
                scan_data_path = mov_path # for export meta data
                thumb_name = os.path.splitext(seq.name)[0] + ".jpg"
                thumb_path = os.path.join(thumbnails_dir, thumb_name)
                print(f"thumb path : {thumb_path} ")
                if mov_to_jpg(scan_data_path, thumb_path):
                    print(f"[OK] Thumbnail created: {thumb_path}")
                else:
                    print(f"[FAIL] Thumbnail create failed: {scan_data_path}")
                    thumb_path = ""

            else:
                print(f"[SKIP] {seq} is not EXR OR JPG")
                continue

            # 메타데이터 추출
            result = subprocess.run(
                ["exiftool", "-json", scan_data_path],
                capture_output=True, 
                text=True
            )

            # meta data json 파싱
            meta = json.loads(result.stdout)[0]
            meta["thumbnail_path"] = thumb_path
            meta["thumbnail"] = thumb_path
            meta_list.append(meta)

    if not meta_list:
        QMessageBox.warning(None, "Load Stopped", "No shots for exporting excel file")
        return None

    # dictionary로 이루어진 list -> xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "Metadata"

    default_fields = ["type", "version", "shot_name", "seq_name", "roll", "thumbnail_path", "thumbnail"]
    all_keys_set = set()
    for m in meta_list:
        for key in m.keys():
            all_keys_set.add(key)

    # 그냥 순서 상관없이 붙이기
    all_fields = default_fields + list(all_keys_set)
    
    # 헤더 작성
    ws.append(all_fields)

    # 데이터 작성
    for row_idx, meta in enumerate(meta_list, start=2):  # 2부터 시작 (1행은 헤더)
        for col_idx, field in enumerate(all_fields, start=1): # 1부터 시작
            value = meta.get(field, "")
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            ws.cell(row=row_idx, column=col_idx, value=value)

        thumb_path = meta.get("thumbnail")
        if thumb_path and os.path.exists(thumb_path):
            col_letter = get_column_letter(all_fields.index("thumbnail") + 1)
            cell_ref = f"{col_letter}{row_idx}"
            try:
                img = Image(thumb_path)
                img.width = 192  
                img.height = 108
                ws.add_image(img, cell_ref)
            except Exception as e:
                print(f"[WARN] Thumbnail insert failed: {e}")

    date = os.path.basename(os.path.normpath(date_path))
    xlsx_path = os.path.join(date_path, f"{date}_meta_output.xlsx")

    wb.save(xlsx_path)

    QMessageBox.information(None, "Export Complete", f"Metadata exported to:\n{xlsx_path}")
    print(f"[COMPLETE] Metadata exported to:\n{xlsx_path}")
    return xlsx_path